# from ecommerce.shipping import calc_shipping
# calc_shipping()

# numbers = [1, 1, 1, 1, 5]
# for x_count in numbers:
#     output = ''
#     for count in range(x_count):
#         output += 'x'
#     print(output)

# names = ['John', 'Bob', 'Mosh', 'Sarah', 'Mary']
# print(names[::3])

# numbers = [3, 2, 4, 5, 6, 7, 5, 5, 6, 8, 7, 9, 10, ]
# max = numbers[0]
# for number in numbers:
#     if number > max:
#         max = number
# print(max)

# matrix = [
#     [1, 3, 4],
#     [2, 4, 2],
#     [7, 8, 9]
# ]
# for row in matrix:
#     for item in row:
#         print(row, item)

# Duplicates
# numbers = [5, 2, 1, 7, 4, 5, 2, 3, 2, 2]
# uniques = []
# for greet in numbers:
#     if greet not in uniques:
#         uniques.append(greet)
# print(uniques)

# # tuples are not modifiable
# numbers = (1, 2, 3)

# # unpacking
# coordinates = [1, 2, 3]
# x, y, z = coordinates
# print(x)

# customer = {
#     "name": "Ayoola",
#     "age": 30,
#     "is_verified": True
# }
# print(customer.get("name"))
# print(customer.get("dob", "Jan 1 1990"))

# phone = input("Phone")
# digits_mapping = {
#     "1": "One",
#     "2": "Two",
#     "3": "Three",
#     "4": "Four"
# }
# output = ""
# for ch in phone:
#     output += digits_mapping.get(ch, "!") + " "
# print(output)

# class Point:
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#
#     def move(self):
#         print("move")
#
#     def draw(self):
#         print("draw")
#
#
# point1 = Point(10,20)
# point1.draw()
# print(point1.x)

# class Person:
#     def __init__(self, name):
#         self.name = name
#
#     def talk(self):
#         print(f"Hi, I am {self.name}")
#
#
# person1 = Person("Ayoola")
# person1.talk()
# # Inheritance
# class Animal:
#     def walk(self):
#         print("walk")
#
#
# class Dog(Animal):
#     pass
#
#
# class Cat(Animal):
#     pass
#
#
# dog1 = Dog()
# cat1 = Cat()
# # Modules
# import converters
# from converters import kg_to_lbs
#
# kg_to_lbs(100)
#
# print(converters.kg_to_lbs(40))


# from utils import find_max
#
# numbers = [10, 3, 6, 2]
# highest = find_max(numbers)
# print(highest)

# import random


# members = ['John', 'Mary', 'Bob', 'Josh']
# leader = random.choice(members)
# print(leader)
#
# class Dice:
#     def roll(self):
#         first = random.randint(1, 6)
#         second = random.randint(1, 6)
#         return first, second
#
#
# dice = Dice()
# print(dice.roll())

# from pathlib import Path
#
# # Absolute Path ==> Starting from Root of hard Disk
# # c:\Program Files\Microsoft ==>Windows
# # /usr/local/bin ==> mac
# # Relative Path ==> Starting from the current dir
#
# path = Path()
# for file in path.glob('*'):
#     print(file)

# working with xl

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


